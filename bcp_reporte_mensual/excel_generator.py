import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows




def generate_xlsx(df:pd.DataFrame, b1=None, b2=None, b3=None):

    # Crear un nuevo Workbook y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active

    # Establecer los valores en las celdas A1:A3 y B1:B3
    ws['A1'] = "Cuenta"
    ws['A2'] = "Moneda"
    ws['A3'] = "Tipo de Cuenta"
    ws['B1'] = b1
    ws['B2'] = b2
    ws['B3'] = b3

    # Estilos para A1:A3
    font_arial_10_bold = Font(name='Arial', size=10, bold=True)
    fill_orange = PatternFill(start_color="FFDFAF", end_color="FFDFAF", fill_type="solid")
    alignment_center = Alignment(vertical="center")

    for row in ws['A1:A3']:
        for cell in row:
            cell.font = font_arial_10_bold
            cell.fill = fill_orange

    for row in ws['A1:B3']:
        for cell in row:
            cell.alignment = alignment_center


    # Establecer el height de las filas 1 a 3
    for i in range(1, 4):
        ws.row_dimensions[i].height = 28.5

    # Establecer el height de la fila 4
    ws.row_dimensions[4].height = 15

    # Añadir el DataFrame a partir de la fila 5
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 5):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Estilos para el header del DataFrame (fila 5)
    fill_blue = PatternFill(start_color="003E9F", end_color="003E9F", fill_type="solid")
    font_white_bold = Font(name='Arial', size=10, color="FFFFFF", bold=True)
    white_border = Border(left=Side(style='thin', color='FFFFFF'),
                        right=Side(style='thin', color='FFFFFF'),
                        top=Side(style='thin', color='FFFFFF'),
                        bottom=Side(style='thin', color='FFFFFF'))

    for cell in ws[5]:
        cell.fill = fill_blue
        cell.font = font_white_bold
        cell.border = white_border
        cell.alignment = alignment_center

    # Establecer el height de la fila 5
    ws.row_dimensions[5].height = 38.5

    # Estilos para las filas del DataFrame (fila 6 en adelante)
    thin_border = Border(left=Side(style='thin', color='C0C0C0'),
                        right=Side(style='thin', color='C0C0C0'),
                        top=Side(style='thin', color='C0C0C0'),
                        bottom=Side(style='thin', color='C0C0C0'))
    fill_even = PatternFill(start_color="F4F4FC", end_color="F4F4FC", fill_type="solid")

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            cell.alignment = alignment_center
            if cell.row % 2 == 1:
                cell.fill = fill_even

    # Ajustar el height de las filas desde la fila 6 en adelante
    for i in range(6, ws.max_row + 1):
        ws.row_dimensions[i].height = 26

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = (24.29**2)/23.57

    ws.sheet_view.showGridLines = False

    # Guardar el archivo Excel
    wb.save("output.xlsx")